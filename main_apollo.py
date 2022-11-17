#Scrape Apollo with selenium
# chrome.exe --remote-debugging-port=8989 --user-data-dir="<directory>"
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import datetime, json, os, subprocess
# import gspread
from time import sleep
from openpyxl import load_workbook


# with open('userinfo.json', 'r') as openfile:
#     SearchLoc = json.load(openfile)
#     openfile.close()

ser = Service("C:\ztool\chromedriver")
op = webdriver.ChromeOptions()
op.add_experimental_option('debuggerAddress', 'localhost:8989')
driver = webdriver.Chrome(service=ser, options=op)

# driver.get('https://app.apollo.io/#/login')
# sleep(2)
# sleep(2)
# driver.get('https://app.apollo.io/')


class OBJ():
    Name = ""
    Linkedin = ""
    Company = ""
    Comurl = ""
    Jobtitle = ""
    Location = ""


listOBJ = []
lst = []
pre_lst = []

while True:
    def dataextract():
        # print(soup.prettify())


        for tr in table.find_all('tbody'): #Enter table tag
            trs = tr.find_all('tr')[-1]
            # print(trs)

            for td in tr:
                tds = td.find_all('td')
                try:
                    obj = OBJ()

                    obj.Linkedin = tds[0].find('div', class_='zp_33Rq5').span.a.get('href') #linkedin url
                    obj.Name = tds[0].text #person name
                    obj.Jobtitle = tds[1].text #jobtitle
                    obj.Company = tds[2].text #company name
                    obj.Comurl = tds[2].find_all('a')[1].get('href').replace("http://www.","") #company url
                    obj.Location = tds[4].text #location
                    obj.Emp = tds[5].text  # employ
                    obj.Email = tds[6].text  # email
                    obj.Industry = tds[7].text  # Industry
                    obj.Keyword = tds[8].text  # Keyword
                    listOBJ.append(obj)

                    lst.append(tds[0].text)

                except:
                    print("There is no data!")

        print("Total number of data " + str(len(listOBJ)))

        for data in listOBJ:
            print(data.Name + " | " + data.Linkedin + " | " + data.Jobtitle + " | " + data.Comurl + " | " + data.Company + " | " + data.Location)

        wb = load_workbook(r'apollo.xlsx')

        sheet = wb.active
        print('Excel Active')

        rows = sheet.max_row
        c1 = rows  # total number of rows
        start = c1 + 1
        end = c1 + 1 + len(listOBJ)
        i = 0
        for data in range(start, end): #data writing
            try:
                name = listOBJ[i].Name
                sheet.cell(row=data, column=1).value = name.split()[0]
                sheet.cell(row=data, column=2).value = name.split()[-1]
                sheet.cell(row=data, column=3).value = listOBJ[i].Comurl
                sheet.cell(row=data, column=4).value = listOBJ[i].Linkedin
                sheet.cell(row=data, column=5).value = listOBJ[i].Company
                sheet.cell(row=data, column=6).value = listOBJ[i].Jobtitle
                sheet.cell(row=data, column=7).value = listOBJ[i].Location
                sheet.cell(row=data, column=8).value = listOBJ[i].Emp
                sheet.cell(row=data, column=9).value = listOBJ[i].Email
                sheet.cell(row=data, column=10).value = listOBJ[i].Industry
                sheet.cell(row=data, column=11).value = listOBJ[i].Keyword

                i += 1
            except:
                pass
        print("Finish!!!\n")
        listOBJ.clear()
        wb.save(r'apollo.xlsx')

        sleep(1)
        next = driver.find_element(By.XPATH,
                                   '//*[@id="provider-mounter"]/div/div[2]/div[2]/div/div[1]/div/div[2]/div/div[2]/div/div/div/div/div[2]/div/div[4]/div/div/div/div/div[3]/div/div[2]/button[2]')
        next.click()

    if __name__ == "__main__":
        inp = int(input("Enter 1 for continue..."))

        if inp ==  1:
            for _ in range(1,101):
                sleep(2)
                html_doc = driver.page_source
                # print(html_doc)
                soup = BeautifulSoup(html_doc, 'html.parser')
                table = soup.find('table')
                # print(table)

                pre_lst = lst.copy()
                lst.clear()

                # print(pre_lst)
                # print(lst)
                # print("before....")

                dataextract() #data extract..................
                if pre_lst == lst:
                    break
                # else:
                # print(pre_lst)
                # print(lst)

        else:
            print("print 1 next time for continue the task...")
